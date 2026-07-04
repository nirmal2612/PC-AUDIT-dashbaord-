# Fanvil Auto Fetch GUI - Python Tkinter
# Date: 03-07-2026
# Purpose: Scan LAN segment, detect Fanvil web phones, fetch MAC/IP and try to fetch web page details after login.
# Safe note: This tool only uses credentials entered by you. It does not bypass login/security.

import os
import re
import csv
import time
import queue
import socket
import threading
import subprocess
from datetime import datetime
from urllib.parse import urljoin

try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
except Exception as e:
    raise SystemExit("Tkinter is required to run this GUI app.")

# Optional packages. Script can still scan without them, but login/detail fetch needs requests + bs4.
try:
    import requests
except Exception:
    requests = None

try:
    from bs4 import BeautifulSoup
except Exception:
    BeautifulSoup = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

OUTPUT_DIR = r"D:\fanvil"
DEFAULT_TIMEOUT = 3

COMMON_INFO_PATHS = [
    "/",
    "/index.htm",
    "/index.html",
    "/cgi-bin/cgiServer.exx",
    "/cgi-bin/ConfigManApp.com",
    "/cgi-bin/ConfigManApp.com?key=STATUS",
    "/cgi-bin/ConfigManApp.com?key=SIP",
    "/cgi-bin/ConfigManApp.com?key=NETWORK",
    "/cgi-bin/ConfigManApp.com?key=DEVICE",
    "/cgi-bin/ConfigManApp.com?key=LINE",
    "/cgi-bin/ConfigManApp.com?key=ACCOUNT",
    "/Status.htm",
    "/status.htm",
    "/phone_status.htm",
    "/System.htm",
    "/system.htm",
    "/Line.htm",
    "/line.htm",
    "/SIP.htm",
    "/sip.htm",
    "/Account.htm",
    "/account.htm",
    "/Network.htm",
    "/network.htm",
]

LOGIN_PATHS = [
    "/",
    "/index.htm",
    "/index.html",
    "/login.htm",
    "/login.html",
    "/cgi-bin/ConfigManApp.com",
]

LOGIN_FIELD_CANDIDATES = [
    ("username", "password"),
    ("user", "password"),
    ("UserName", "Password"),
    ("loginname", "password"),
    ("account", "password"),
    ("name", "pwd"),
    ("username", "pwd"),
    ("admin", "password"),
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) FanvilAutoFetch/1.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

EXPORT_COLUMNS = [
    "S.No", "IP Address", "MAC Address", "Brand", "Model", "Firmware", "Device Name",
    "Extension", "SIP Username", "Display Name", "Phone Number", "Line Status",
    "Web Status", "Login Status", "Matched Source", "Remarks", "Scan Time"
]


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def ping_ip(ip: str, timeout_ms: int = 800) -> bool:
    try:
        result = subprocess.run(
            ["ping", "-n", "1", "-w", str(timeout_ms), ip],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
        )
        return result.returncode == 0
    except Exception:
        return False


def get_arp_mac(ip: str) -> str:
    try:
        # Ping first so ARP table refreshes.
        ping_ip(ip, 500)
        result = subprocess.run(
            ["arp", "-a", ip],
            capture_output=True,
            text=True,
            timeout=3,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0,
        )
        text = result.stdout + "\n" + result.stderr
        # Windows ARP MAC format: xx-xx-xx-xx-xx-xx
        m = re.search(r"([0-9a-fA-F]{2}[-:]){5}[0-9a-fA-F]{2}", text)
        if m:
            return m.group(0).upper().replace("-", ":")
    except Exception:
        pass
    return ""


def check_tcp_port(ip: str, port: int, timeout: float = 1.5) -> bool:
    try:
        with socket.create_connection((ip, port), timeout=timeout):
            return True
    except Exception:
        return False


def safe_text(html: str) -> str:
    if not html:
        return ""
    if BeautifulSoup:
        try:
            soup = BeautifulSoup(html, "html.parser")
            for tag in soup(["script", "style"]):
                tag.extract()
            return soup.get_text(" ", strip=True)
        except Exception:
            pass
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html)).strip()


def looks_fanvil(text: str, headers_text: str = "") -> bool:
    joined = (text + " " + headers_text).lower()
    keywords = ["fanvil", "x301", "x3s", "x3sg", "x4", "x5", "x6", "voip phone", "sip phone"]
    return any(k in joined for k in keywords)


def find_value_by_patterns(text: str, patterns):
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            val = m.group(1).strip()
            val = re.sub(r"[\s\|,;]+$", "", val)
            if val and len(val) < 120:
                return val
    return ""


def parse_details_from_text(text: str):
    details = {}
    details["Model"] = find_value_by_patterns(text, [
        r"Model(?:\s*Number)?\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
        r"Product(?:\s*Name)?\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
        r"Device\s*Model\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
    ])
    details["Firmware"] = find_value_by_patterns(text, [
        r"Firmware(?:\s*Version)?\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
        r"Software(?:\s*Version)?\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
        r"Version\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
    ])
    details["Device Name"] = find_value_by_patterns(text, [
        r"Device\s*Name\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
        r"Phone\s*Name\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
    ])
    details["Extension"] = find_value_by_patterns(text, [
        r"Extension\s*[:=]\s*([0-9A-Za-z_\-\.]+)",
        r"Ext\.?\s*[:=]\s*([0-9A-Za-z_\-\.]+)",
        r"Line\s*1\s*(?:User|Account|Extension)?\s*[:=]\s*([0-9A-Za-z_\-\.]+)",
    ])
    details["SIP Username"] = find_value_by_patterns(text, [
        r"SIP\s*User(?:\s*Name)?\s*[:=]\s*([0-9A-Za-z_\-\.@]+)",
        r"Auth(?:entication)?\s*User(?:\s*Name)?\s*[:=]\s*([0-9A-Za-z_\-\.@]+)",
        r"Register\s*Name\s*[:=]\s*([0-9A-Za-z_\-\.@]+)",
        r"User\s*Name\s*[:=]\s*([0-9A-Za-z_\-\.@]+)",
    ])
    details["Display Name"] = find_value_by_patterns(text, [
        r"Display\s*Name\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
        r"Label\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
    ])
    details["Phone Number"] = find_value_by_patterns(text, [
        r"Phone\s*Number\s*[:=]\s*([0-9A-Za-z_\-\.]+)",
        r"Number\s*[:=]\s*([0-9]{3,15})",
    ])
    details["Line Status"] = find_value_by_patterns(text, [
        r"Line\s*Status\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
        r"Register(?:ed|\s*Status)?\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
        r"Registration\s*Status\s*[:=]\s*([A-Za-z0-9_\-\. ]+)",
    ])
    return details


def parse_details_from_inputs(html: str):
    details = {}
    if not BeautifulSoup or not html:
        return details
    try:
        soup = BeautifulSoup(html, "html.parser")
        # Inputs/selects often contain actual configuration values.
        candidates = {}
        for tag in soup.find_all(["input", "select", "textarea"]):
            name = (tag.get("name") or tag.get("id") or "").strip()
            value = (tag.get("value") or tag.text or "").strip()
            if name and value:
                candidates[name.lower()] = value

        key_map = {
            "Model": ["model", "phone_model", "devicemodel", "device_model", "product"],
            "Firmware": ["firmware", "fwversion", "firmwareversion", "version", "software"],
            "Device Name": ["devicename", "device_name", "phone_name", "hostname"],
            "Extension": ["extension", "ext", "line1_extension", "account1_extension", "sip1_extension"],
            "SIP Username": ["sip user", "sipuser", "sip_user", "username", "user_name", "authuser", "auth_user", "registername", "register_name", "account"],
            "Display Name": ["displayname", "display_name", "label", "line1_displayname"],
            "Phone Number": ["phonenumber", "phone_number", "number", "telnum"],
            "Line Status": ["linestatus", "line_status", "registerstatus", "register_status", "registration"],
        }
        for out_key, names in key_map.items():
            if details.get(out_key):
                continue
            for cname, cval in candidates.items():
                if any(n in cname for n in names):
                    if cval and len(cval) < 120:
                        details[out_key] = cval
                        break
    except Exception:
        pass
    return details


def merge_details(base, extra):
    for k, v in extra.items():
        if v and not base.get(k):
            base[k] = v
    return base


def try_generic_login(session, base_url, username, password, timeout):
    """Generic login handler. Fanvil firmware differs, so this tries common forms and common field names."""
    if not username:
        return False, "No username entered"

    for path in LOGIN_PATHS:
        url = urljoin(base_url, path)
        try:
            r = session.get(url, headers=HEADERS, timeout=timeout, verify=False)
            html = r.text or ""
            text = safe_text(html)
            # Basic auth fallback: requests auth is tried separately in caller if needed.
            if BeautifulSoup and html:
                soup = BeautifulSoup(html, "html.parser")
                forms = soup.find_all("form")
                for form in forms or [None]:
                    action = form.get("action") if form else path
                    method = (form.get("method") or "post").lower() if form else "post"
                    post_url = urljoin(url, action or path)

                    # Preserve hidden fields/tokens.
                    payload = {}
                    if form:
                        for inp in form.find_all(["input", "select", "textarea"]):
                            nm = inp.get("name")
                            if nm:
                                payload[nm] = inp.get("value", "")

                    for ufield, pfield in LOGIN_FIELD_CANDIDATES:
                        data = dict(payload)
                        data[ufield] = username
                        data[pfield] = password
                        try:
                            if method == "get":
                                rr = session.get(post_url, params=data, headers=HEADERS, timeout=timeout, verify=False)
                            else:
                                rr = session.post(post_url, data=data, headers=HEADERS, timeout=timeout, verify=False)
                            combined = safe_text(rr.text)
                            low = combined.lower()
                            if rr.status_code in (200, 302, 301) and not any(x in low for x in ["login failed", "invalid password", "password error", "incorrect"]):
                                # Verify by checking a protected-ish status page or absence of login form.
                                if "logout" in low or "status" in low or "account" in low or looks_fanvil(combined, str(rr.headers)):
                                    return True, f"Login attempted via {path} using {ufield}/{pfield}"
                        except Exception:
                            continue

            # Some devices accept direct POST without obvious form.
            for ufield, pfield in LOGIN_FIELD_CANDIDATES:
                data = {ufield: username, pfield: password}
                try:
                    rr = session.post(url, data=data, headers=HEADERS, timeout=timeout, verify=False)
                    low = safe_text(rr.text).lower()
                    if rr.status_code in (200, 302, 301) and ("logout" in low or "status" in low or "account" in low):
                        return True, f"Direct POST via {path} using {ufield}/{pfield}"
                except Exception:
                    pass
        except Exception:
            continue
    return False, "Login not confirmed"


def fetch_fanvil_details(ip: str, username: str, password: str, timeout: int = DEFAULT_TIMEOUT):
    row = {
        "IP Address": ip,
        "MAC Address": get_arp_mac(ip),
        "Brand": "",
        "Model": "",
        "Firmware": "",
        "Device Name": "",
        "Extension": "",
        "SIP Username": "",
        "Display Name": "",
        "Phone Number": "",
        "Line Status": "",
        "Web Status": "Not checked",
        "Login Status": "Not attempted",
        "Matched Source": "",
        "Remarks": "",
        "Scan Time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    if requests is None:
        row["Remarks"] = "Install required package: pip install requests beautifulsoup4 openpyxl"
        return row

    ports = []
    for p in [80, 8080, 443]:
        if check_tcp_port(ip, p):
            ports.append(p)
    if not ports:
        row["Web Status"] = "No HTTP/HTTPS port"
        row["Remarks"] = "Device pinged but no common web port open"
        return row

    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    base_urls = []
    for p in ports:
        if p == 443:
            base_urls.append(f"https://{ip}/")
        elif p == 80:
            base_urls.append(f"http://{ip}/")
        else:
            base_urls.append(f"http://{ip}:{p}/")

    for base_url in base_urls:
        session = requests.Session()
        session.headers.update(HEADERS)
        try:
            r = session.get(base_url, timeout=timeout, verify=False)
            text = safe_text(r.text)
            header_text = str(r.headers)
            row["Web Status"] = f"HTTP {r.status_code}"
            if looks_fanvil(text, header_text):
                row["Brand"] = "Fanvil"
                row["Matched Source"] = base_url
                merge_details(row, parse_details_from_text(text))
                merge_details(row, parse_details_from_inputs(r.text))
            elif not row["Brand"]:
                row["Remarks"] = "Web page found, Fanvil not confirmed"

            # Try Basic Auth quick check as some devices/pages may allow it.
            login_ok = False
            login_msg = "Login not confirmed"
            if username:
                login_ok, login_msg = try_generic_login(session, base_url, username, password, timeout)
                if not login_ok:
                    try:
                        br = requests.get(base_url, auth=(username, password), headers=HEADERS, timeout=timeout, verify=False)
                        if br.status_code == 200 and looks_fanvil(safe_text(br.text), str(br.headers)):
                            session.auth = (username, password)
                            login_ok = True
                            login_msg = "Basic Auth accepted"
                    except Exception:
                        pass
                row["Login Status"] = "Success" if login_ok else login_msg
            else:
                row["Login Status"] = "Skipped - no username"

            # Fetch common status/config paths. Even if login isn't confirmed, some details may be public.
            found_any = False
            for path in COMMON_INFO_PATHS:
                url = urljoin(base_url, path)
                try:
                    rr = session.get(url, timeout=timeout, verify=False)
                    if rr.status_code not in (200, 401, 403):
                        continue
                    body_text = safe_text(rr.text)
                    if rr.status_code == 200:
                        if looks_fanvil(body_text, str(rr.headers)):
                            row["Brand"] = "Fanvil"
                            if not row["Matched Source"]:
                                row["Matched Source"] = url
                        before = dict(row)
                        merge_details(row, parse_details_from_text(body_text))
                        merge_details(row, parse_details_from_inputs(rr.text))
                        if dict(row) != before:
                            found_any = True
                            if not row["Matched Source"]:
                                row["Matched Source"] = url
                    elif rr.status_code in (401, 403) and not row["Remarks"]:
                        row["Remarks"] = "Protected page detected"
                except Exception:
                    continue

            if row["Brand"] == "Fanvil" or found_any:
                if not row["Remarks"]:
                    if row["Login Status"] == "Success":
                        row["Remarks"] = "Details fetched where fields were available"
                    else:
                        row["Remarks"] = "Fanvil detected; login/details may need model-specific parser"
                return row

        except Exception as e:
            row["Web Status"] = "Web error"
            row["Remarks"] = str(e)[:180]
            continue

    return row


def save_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=EXPORT_COLUMNS)
        writer.writeheader()
        for i, row in enumerate(rows, start=1):
            out = {col: row.get(col, "") for col in EXPORT_COLUMNS}
            out["S.No"] = i
            writer.writerow(out)


def save_xlsx(rows, path):
    if Workbook is None:
        csv_path = os.path.splitext(path)[0] + ".csv"
        save_csv(rows, csv_path)
        return csv_path

    wb = Workbook()
    ws = wb.active
    ws.title = "Fanvil Scan"
    ws.append(EXPORT_COLUMNS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for i, row in enumerate(rows, start=1):
        out = {col: row.get(col, "") for col in EXPORT_COLUMNS}
        out["S.No"] = i
        ws.append([out[col] for col in EXPORT_COLUMNS])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col_idx, col_name in enumerate(EXPORT_COLUMNS, start=1):
        width = max(len(col_name) + 2, 14)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            val = str(row[0].value or "")
            width = min(max(width, len(val) + 2), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
    return path


class FanvilApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Fanvil Auto Fetch Tool")
        self.root.geometry("1180x680")
        self.root.minsize(1000, 600)

        self.results = []
        self.log_queue = queue.Queue()
        self.stop_flag = threading.Event()
        self.worker_thread = None

        self.build_ui()
        self.root.after(200, self.process_log_queue)

    def build_ui(self):
        top = ttk.Frame(self.root, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="Fanvil Auto Fetch Tool", font=("Segoe UI", 16, "bold")).grid(row=0, column=0, columnspan=8, sticky="w", pady=(0, 8))

        ttk.Label(top, text="IP Segment").grid(row=1, column=0, sticky="w")
        self.segment_var = tk.StringVar(value="10.209.110")
        ttk.Entry(top, textvariable=self.segment_var, width=18).grid(row=1, column=1, padx=5, sticky="w")
        ttk.Label(top, text="Start").grid(row=1, column=2, sticky="w")
        self.start_var = tk.IntVar(value=1)
        ttk.Entry(top, textvariable=self.start_var, width=6).grid(row=1, column=3, padx=5)
        ttk.Label(top, text="End").grid(row=1, column=4, sticky="w")
        self.end_var = tk.IntVar(value=254)
        ttk.Entry(top, textvariable=self.end_var, width=6).grid(row=1, column=5, padx=5)

        ttk.Label(top, text="Username").grid(row=2, column=0, sticky="w", pady=(8, 0))
        self.user_var = tk.StringVar(value="admin")
        ttk.Entry(top, textvariable=self.user_var, width=18).grid(row=2, column=1, padx=5, sticky="w", pady=(8, 0))
        ttk.Label(top, text="Password").grid(row=2, column=2, sticky="w", pady=(8, 0))
        self.pass_var = tk.StringVar(value="admin")
        ttk.Entry(top, textvariable=self.pass_var, width=18, show="*").grid(row=2, column=3, padx=5, sticky="w", pady=(8, 0))
        ttk.Label(top, text="Threads").grid(row=2, column=4, sticky="w", pady=(8, 0))
        self.thread_var = tk.IntVar(value=20)
        ttk.Entry(top, textvariable=self.thread_var, width=6).grid(row=2, column=5, padx=5, pady=(8, 0))

        self.scan_btn = ttk.Button(top, text="SCAN & EXPORT", command=self.start_scan)
        self.scan_btn.grid(row=1, column=6, rowspan=2, padx=10, sticky="ns")
        self.stop_btn = ttk.Button(top, text="STOP", command=self.stop_scan, state="disabled")
        self.stop_btn.grid(row=1, column=7, rowspan=2, sticky="ns")

        self.progress = ttk.Progressbar(self.root, mode="determinate")
        self.progress.pack(fill="x", padx=10, pady=(0, 8))

        columns = EXPORT_COLUMNS
        table_frame = ttk.Frame(self.root, padding=(10, 0, 10, 5))
        table_frame.pack(fill="both", expand=True)
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=130, anchor="w")
        self.tree.column("S.No", width=55, anchor="center")
        self.tree.column("IP Address", width=120)
        self.tree.column("MAC Address", width=150)
        self.tree.column("Remarks", width=260)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        bottom = ttk.Frame(self.root, padding=10)
        bottom.pack(fill="x")
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(bottom, textvariable=self.status_var).pack(side="left")
        ttk.Button(bottom, text="Open Output Folder", command=self.open_output_folder).pack(side="right")

    def log(self, msg):
        self.log_queue.put(("status", msg))

    def add_row(self, row):
        self.log_queue.put(("row", row))

    def process_log_queue(self):
        try:
            while True:
                typ, payload = self.log_queue.get_nowait()
                if typ == "status":
                    self.status_var.set(payload)
                elif typ == "progress":
                    self.progress["value"] = payload
                elif typ == "row":
                    self.results.append(payload)
                    i = len(self.results)
                    values = []
                    for col in EXPORT_COLUMNS:
                        values.append(i if col == "S.No" else payload.get(col, ""))
                    self.tree.insert("", "end", values=values)
        except queue.Empty:
            pass
        self.root.after(200, self.process_log_queue)

    def open_output_folder(self):
        ensure_output_dir()
        try:
            os.startfile(OUTPUT_DIR)
        except Exception:
            messagebox.showinfo("Output Folder", OUTPUT_DIR)

    def validate_inputs(self):
        seg = self.segment_var.get().strip()
        if not re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}$", seg):
            messagebox.showerror("Invalid Segment", "Enter segment like 10.209.110")
            return None
        nums = [int(x) for x in seg.split(".")]
        if any(n < 0 or n > 255 for n in nums):
            messagebox.showerror("Invalid Segment", "Segment numbers must be 0-255")
            return None
        start = int(self.start_var.get())
        end = int(self.end_var.get())
        if start < 1 or end > 254 or start > end:
            messagebox.showerror("Invalid Range", "Start/End must be 1-254 and Start <= End")
            return None
        threads = int(self.thread_var.get())
        threads = max(1, min(threads, 80))
        return seg, start, end, threads

    def start_scan(self):
        if self.worker_thread and self.worker_thread.is_alive():
            return
        valid = self.validate_inputs()
        if not valid:
            return
        if requests is None or BeautifulSoup is None:
            messagebox.showwarning(
                "Missing Packages",
                "For login/detail fetching, install packages:\n\npython -m pip install requests beautifulsoup4 openpyxl\n\nScanning can still run, but details may be limited."
            )
        self.results.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.progress["value"] = 0
        self.stop_flag.clear()
        self.scan_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.worker_thread = threading.Thread(target=self.scan_worker, args=valid, daemon=True)
        self.worker_thread.start()

    def stop_scan(self):
        self.stop_flag.set()
        self.status_var.set("Stopping scan...")

    def scan_worker(self, valid):
        seg, start, end, threads = valid
        username = self.user_var.get().strip()
        password = self.pass_var.get()
        ips = [f"{seg}.{i}" for i in range(start, end + 1)]
        total = len(ips)
        completed = 0
        lock = threading.Lock()
        q_ips = queue.Queue()
        for ip in ips:
            q_ips.put(ip)

        self.log(f"Scanning {total} IPs...")

        def worker():
            nonlocal completed
            while not q_ips.empty() and not self.stop_flag.is_set():
                try:
                    ip = q_ips.get_nowait()
                except queue.Empty:
                    break
                try:
                    # Ping first to skip dead devices faster.
                    alive = ping_ip(ip, 700)
                    web_open = check_tcp_port(ip, 80) or check_tcp_port(ip, 8080) or check_tcp_port(ip, 443)
                    if alive or web_open:
                        row = fetch_fanvil_details(ip, username, password, DEFAULT_TIMEOUT)
                        # Add all web devices, but Fanvil gets Brand marked. This helps see non-Fanvil pages too.
                        if row.get("Brand") == "Fanvil" or row.get("Web Status") not in ["No HTTP/HTTPS port", "Not checked"]:
                            self.add_row(row)
                except Exception as e:
                    self.add_row({
                        "IP Address": ip,
                        "MAC Address": get_arp_mac(ip),
                        "Web Status": "Error",
                        "Login Status": "Error",
                        "Remarks": str(e)[:180],
                        "Scan Time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    })
                finally:
                    with lock:
                        completed += 1
                        self.log_queue.put(("progress", int((completed / total) * 100)))
                        self.log(f"Scanned {completed}/{total} | Found rows: {len(self.results)}")
                    q_ips.task_done()

        workers = []
        for _ in range(threads):
            t = threading.Thread(target=worker, daemon=True)
            workers.append(t)
            t.start()
        for t in workers:
            t.join()

        ensure_output_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        xlsx_path = os.path.join(OUTPUT_DIR, f"Fanvil_AutoFetch_{timestamp}.xlsx")
        saved_path = save_xlsx(self.results, xlsx_path)
        self.log(f"Completed. Saved: {saved_path}")
        self.root.after(0, lambda: self.scan_btn.config(state="normal"))
        self.root.after(0, lambda: self.stop_btn.config(state="disabled"))
        self.root.after(0, lambda: messagebox.showinfo("Scan Completed", f"Rows: {len(self.results)}\nSaved:\n{saved_path}"))


def main():
    ensure_output_dir()
    root = tk.Tk()
    app = FanvilApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
